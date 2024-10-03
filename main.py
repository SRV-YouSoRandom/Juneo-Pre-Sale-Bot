import discord
from discord import app_commands
from discord.ext import commands
import openpyxl
import os
from dotenv import load_dotenv

# Load environment variables from .env file
load_dotenv()
TOKEN = os.getenv("DISCORD_TOKEN")

# Set intents
intents = discord.Intents.default()
intents.message_content = True
intents.guilds = True

# Bot initialization
bot = commands.Bot(command_prefix="!", intents=intents)

# Create the Excel sheet if it doesn't exist
def create_excel_file():
    if not os.path.exists("user_data.xlsx"):
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.append(["Discord Username", "User ID", "Wallet Address", "Role"])
        workbook.save("user_data.xlsx")

# Permission check decorator
def has_required_role():
    async def predicate(interaction: discord.Interaction):
        role_names = [role.name for role in interaction.user.roles]
        print(f"User {interaction.user.name} has roles: {role_names}")  # Debugging

        if "Admin" in role_names or "Tech Mod" in role_names:
            return True
        await interaction.response.send_message("You don't have permission to use this command.", ephemeral=True)
        return False
    return app_commands.check(predicate)

# Slash command to setup the bot
@bot.tree.command(name="setup", description="Setup the bot with buttons (Admin/Tech Mod only)")
@has_required_role()
async def setup(interaction: discord.Interaction):
    print(f"User {interaction.user.name} is trying to run /setup command")  # Debugging
    # Create buttons
    view = discord.ui.View()

    # Connect button
    view.add_item(discord.ui.Button(label="Connect", style=discord.ButtonStyle.primary, custom_id="connect_button"))

    # Verify Role button
    view.add_item(discord.ui.Button(label="Verify Role", style=discord.ButtonStyle.success, custom_id="verify_button"))

    # Modify Address button
    view.add_item(discord.ui.Button(label="Modify Address", style=discord.ButtonStyle.secondary, custom_id="modify_button"))

    await interaction.response.send_message("Click a button to link or verify your wallet address.", view=view)

# Slash command to download the Excel sheet
@bot.tree.command(name="download", description="Download the user data (Admin/Tech Mod only)")
@has_required_role()
async def download(interaction: discord.Interaction):
    create_excel_file()
    await interaction.response.send_message(file=discord.File("user_data.xlsx"), ephemeral=True)

# Connect Wallet Modal
class ConnectWalletModal(discord.ui.Modal, title="Connect Your Wallet"):
    wallet_address = discord.ui.TextInput(label="Wallet Address", placeholder="Enter your Web3 wallet address")

    async def on_submit(self, interaction: discord.Interaction):
        create_excel_file()
        user = interaction.user
        role_names = [role.name for role in user.roles]
        special_roles = [role for role in role_names if role in ["Odin", "Titan", "Orbital"]]

        # Load the workbook and check if the user already exists
        workbook = openpyxl.load_workbook("user_data.xlsx")
        sheet = workbook.active
        user_exists = False

        for row in sheet.iter_rows(values_only=True):
            if row[1] == str(user.id):  # Check if the User ID already exists
                user_exists = True
                break

        if user_exists:
            await interaction.response.send_message("You have already connected a wallet. Use 'Modify Address' to change it.", ephemeral=True)
        else:
            # Add new entry to the Excel sheet if the user doesn't exist
            sheet.append([user.name, str(user.id), self.wallet_address.value, ', '.join(special_roles)])
            workbook.save("user_data.xlsx")
            await interaction.response.send_message(f"Wallet address {self.wallet_address.value} connected successfully!", ephemeral=True)

# Handle button interactions
@bot.event
async def on_interaction(interaction: discord.Interaction):
    if interaction.type == discord.InteractionType.component:
        if interaction.data["custom_id"] == "connect_button":
            modal = ConnectWalletModal()
            await interaction.response.send_modal(modal)

        elif interaction.data["custom_id"] == "verify_button":
            create_excel_file()
            workbook = openpyxl.load_workbook("user_data.xlsx")
            sheet = workbook.active
            user = interaction.user
            user_found = False
            for row in sheet.iter_rows(values_only=True):
                if row[1] == str(user.id):
                    user_found = True
                    role = discord.utils.get(user.guild.roles, name="wallet verified")
                    if role:
                        await user.add_roles(role)
                        await interaction.response.send_message(f"Your role has been verified.", ephemeral=True)
                    else:
                        await interaction.response.send_message(f"The 'wallet verified' role does not exist.", ephemeral=True)
            if not user_found:
                await interaction.response.send_message("Please connect your wallet first using the 'Connect' button.", ephemeral=True)

        elif interaction.data["custom_id"] == "modify_button":
            # Check if user has already linked their wallet
            create_excel_file()
            workbook = openpyxl.load_workbook("user_data.xlsx")
            sheet = workbook.active
            user = interaction.user
            user_row = None
            for i, row in enumerate(sheet.iter_rows(values_only=True)):
                if row[1] == str(user.id):
                    user_row = i + 1
                    break

            if user_row:
                # If user has linked a wallet, allow them to modify it
                class ModifyWalletModal(discord.ui.Modal, title="Modify Wallet Address"):
                    new_wallet_address = discord.ui.TextInput(label="New Wallet Address", placeholder="Enter new wallet address")

                    async def on_submit(self, interaction: discord.Interaction):
                        sheet.cell(row=user_row, column=3).value = self.new_wallet_address.value
                        workbook.save("user_data.xlsx")
                        await interaction.response.send_message(f"Your wallet address has been updated to {self.new_wallet_address.value}", ephemeral=True)

                modal = ModifyWalletModal()
                await interaction.response.send_modal(modal)
            else:
                await interaction.response.send_message("You haven't linked a wallet yet. Please connect first.", ephemeral=True)

# Handle bot ready event
@bot.event
async def on_ready():
    await bot.tree.sync()
    print(f"Logged in as {bot.user}")
    create_excel_file()

# Run the bot
bot.run(TOKEN)
